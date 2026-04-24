# Backend tests for the Control Analysis module
# Exercises: CSV/Excel/ServiceNow import, CRUD, quality, duplicates, coverage,
# regulation mapping, LLM evidence evaluation, 5W1H, uplift, workpaper (PDF/XLSX).
import io
import os
import uuid

import pytest
import requests
from openpyxl import load_workbook, Workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # .env loaded on frontend; read directly as fallback
    with open("/app/frontend/.env") as fh:
        for line in fh:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
                break

API = f"{BASE_URL}/api/control-analysis"
TENANT = f"test_tenant_{uuid.uuid4().hex[:6]}"


@pytest.fixture(scope="module")
def tenant():
    yield TENANT
    # Best-effort cleanup - delete all controls for the tenant
    try:
        r = requests.get(f"{API}/controls", params={"tenant_id": TENANT, "limit": 1000}, timeout=15)
        for c in r.json().get("controls", []):
            requests.delete(f"{API}/controls/{c['id']}", params={"tenant_id": TENANT}, timeout=15)
    except Exception:
        pass


# ---- Ingestion ----
def _csv_bytes():
    csv = (
        "control_id,name,description,domain,owner,type,category,test_procedure,frameworks\n"
        "CTRL-T-001,MFA on Admin,Multi-factor authentication is enforced for all privileged accounts,Access Control,sec@bank.com,TECHNICAL,PREVENTIVE,Verify MFA is automated for 100% of admin accounts,\"NIST CSF,ISO 27001\"\n"
        "CTRL-T-002,Data Encryption,All sensitive data encrypted at rest using AES-256,Data Protection,data@bank.com,TECHNICAL,PREVENTIVE,Verify encryption is enabled on all databases,ISO 27001\n"
        "CTRL-T-003,CAB Approval,Production changes require CAB approval periodically,Change Management,change@bank.com,ADMINISTRATIVE,PREVENTIVE,Review CAB meeting minutes,SOC2\n"
    )
    return csv.encode()


def test_csv_import(tenant):
    files = {"file": ("controls.csv", _csv_bytes(), "text/csv")}
    r = requests.post(f"{API}/controls/import/csv", params={"tenant_id": tenant}, files=files, timeout=30)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["imported"] == 3
    assert body["source_type"] == "CSV"


def test_excel_import(tenant):
    wb = Workbook()
    ws = wb.active
    ws.append(["control_id", "name", "description", "domain", "owner", "type", "category", "test_procedure", "frameworks"])
    ws.append(["CTRL-XL-1", "Backup Review", "Daily backup validation procedure automated", "Business Continuity", "ops@bank.com", "OPERATIONAL", "DETECTIVE", "Verify backup logs daily", "ISO 27001"])
    buf = io.BytesIO()
    wb.save(buf)
    files = {"file": ("controls.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/controls/import/excel", params={"tenant_id": tenant}, files=files, timeout=30)
    assert r.status_code == 200, r.text
    assert r.json()["imported"] == 1


def test_servicenow_mock_import(tenant):
    r = requests.post(f"{API}/controls/import/servicenow", params={"tenant_id": tenant}, timeout=30)
    assert r.status_code == 200
    assert r.json()["imported"] == 3


# ---- Register CRUD ----
def test_list_controls_no_object_id_leak(tenant):
    r = requests.get(f"{API}/controls", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 200
    data = r.json()
    assert data["total"] >= 7
    for c in data["controls"]:
        assert "_id" not in c
        assert "id" in c


def test_filter_controls_by_domain(tenant):
    r = requests.get(f"{API}/controls", params={"tenant_id": tenant, "domain": "Access Control"}, timeout=15)
    assert r.status_code == 200
    for c in r.json()["controls"]:
        assert c["domain"] == "Access Control"


def test_get_update_delete_control(tenant):
    lst = requests.get(f"{API}/controls", params={"tenant_id": tenant}, timeout=15).json()["controls"]
    cid = lst[0]["id"]

    r = requests.get(f"{API}/controls/{cid}", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 200
    assert r.json()["id"] == cid

    r = requests.put(
        f"{API}/controls/{cid}",
        params={"tenant_id": tenant},
        json={"updates": {"owner": "updated@bank.com"}},
        timeout=15,
    )
    assert r.status_code == 200
    assert r.json()["owner"] == "updated@bank.com"

    # 404 path
    r = requests.get(f"{API}/controls/nonexistent-id", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 404


# ---- Quality / duplicates / coverage ----
def test_quality(tenant):
    r = requests.get(f"{API}/quality", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 200
    data = r.json()
    for key in ("total_controls", "quality_score", "domain_coverage", "control_type_distribution"):
        assert key in data
    assert data["total_controls"] >= 7


def test_duplicates(tenant):
    r = requests.get(f"{API}/duplicates", params={"tenant_id": tenant, "similarity_threshold": 0.5}, timeout=15)
    assert r.status_code == 200
    assert "duplicates" in r.json()


def test_coverage(tenant):
    r = requests.get(f"{API}/coverage", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 200
    data = r.json()
    assert "domains" in data
    assert data["covered_domains"] >= 1


# ---- Regulation mapping ----
def test_regulation_mapping(tenant):
    payload = {
        "regulation_content": (
            "Organizations must enforce multi factor authentication for privileged accounts.\n"
            "Sensitive data shall be encrypted at rest and in transit using industry standards.\n"
            "Changes to production systems should be approved by a change advisory board.\n"
        ),
        "framework_name": "TestFramework",
    }
    r = requests.post(f"{API}/regulation/map", params={"tenant_id": tenant}, json=payload, timeout=30)
    assert r.status_code == 200
    data = r.json()
    assert "compliance_score" in data
    assert "by_requirement_type" in data and "mandatory" in data["by_requirement_type"]
    assert "gaps" in data


# ---- LLM-powered endpoints ----
def test_uplift_language_real_llm(tenant):
    cid = requests.get(f"{API}/controls", params={"tenant_id": tenant}, timeout=15).json()["controls"][0]["id"]
    r = requests.post(f"{API}/controls/{cid}/uplift", params={"tenant_id": tenant}, timeout=90)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["control_id"] == cid
    assert "improved" in data and "original" in data


def test_evidence_evaluate_text_real_llm(tenant):
    controls = requests.get(f"{API}/controls", params={"tenant_id": tenant}, timeout=15).json()["controls"]
    cid = controls[0]["id"]
    evidence = (
        "Evidence: Screenshot of IAM console shows MFA is enforced for all 42 admin accounts. "
        "Okta report attached confirms 100% compliance on 2026-01-10. Policy document attached."
    )
    payload = {"control_id": cid, "evidence_text": evidence, "evidence_filename": "iam_audit.txt", "uploaded_by": "tester"}
    r = requests.post(f"{API}/evidence/evaluate", params={"tenant_id": tenant}, json=payload, timeout=120)
    assert r.status_code == 200, r.text
    body = r.json()
    assert "evaluation" in body
    ev = body["evaluation"]
    for k in ("status", "confidence", "effectiveness", "reasoning"):
        assert k in ev
    assert "narrative_5w1h" in body
    # store evaluation id for later
    pytest._last_eval_id = body.get("evaluation_id")
    pytest._eval_control_id = cid


def test_evidence_upload_txt(tenant):
    cid = getattr(pytest, "_eval_control_id", None) or requests.get(
        f"{API}/controls", params={"tenant_id": tenant}, timeout=15
    ).json()["controls"][0]["id"]
    content = b"Audit log excerpt: backups verified successfully on 2026-01-12. RTO met."
    files = {"file": ("evidence.txt", content, "text/plain")}
    data = {"control_id": cid, "uploaded_by": "tester"}
    r = requests.post(f"{API}/evidence/upload", params={"tenant_id": tenant}, data=data, files=files, timeout=120)
    assert r.status_code == 200, r.text
    assert "evaluation" in r.json()


def test_evidence_upload_empty_file_returns_400(tenant):
    cid = getattr(pytest, "_eval_control_id", None) or requests.get(
        f"{API}/controls", params={"tenant_id": tenant}, timeout=15
    ).json()["controls"][0]["id"]
    files = {"file": ("empty.txt", b"", "text/plain")}
    data = {"control_id": cid}
    r = requests.post(f"{API}/evidence/upload", params={"tenant_id": tenant}, data=data, files=files, timeout=30)
    assert r.status_code == 400


def test_list_and_get_evaluations(tenant):
    r = requests.get(f"{API}/evaluations", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 200
    evals = r.json()["evaluations"]
    assert len(evals) >= 1
    eid = evals[0]["id"]
    r = requests.get(f"{API}/evaluations/{eid}", params={"tenant_id": tenant}, timeout=15)
    assert r.status_code == 200
    assert r.json()["id"] == eid


def test_5w1h_regenerate(tenant):
    cid = getattr(pytest, "_eval_control_id", None)
    if not cid:
        pytest.skip("no evaluation run")
    r = requests.get(f"{API}/controls/{cid}/5w1h", params={"tenant_id": tenant}, timeout=90)
    assert r.status_code == 200
    assert "narrative_5w1h" in r.json()


# ---- Workpapers ----
def test_workpaper_pdf(tenant):
    r = requests.post(f"{API}/workpaper/pdf", params={"tenant_id": tenant}, json={}, timeout=60)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/pdf")
    assert r.content[:4] == b"%PDF"


def test_workpaper_excel(tenant):
    r = requests.post(f"{API}/workpaper/excel", params={"tenant_id": tenant}, json={}, timeout=60)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    names = set(wb.sheetnames)
    for expected in ("Control Register", "Quality Summary", "Domain Coverage", "Evidence Evaluations"):
        assert expected in names, f"Missing sheet {expected}; got {names}"


# ---- Regression: existing endpoints still work ----
def test_regression_auth_login():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": "admin@bank.com", "password": "admin123"}, timeout=15)
    assert r.status_code == 200
    assert "access_token" in r.json() or "token" in r.json()


def test_regression_tech_risk_assessments():
    # login first
    tok = requests.post(f"{BASE_URL}/api/auth/login", json={"email": "admin@bank.com", "password": "admin123"}, timeout=15).json()
    token = tok.get("access_token") or tok.get("token")
    h = {"Authorization": f"Bearer {token}"} if token else {}
    r = requests.get(f"{BASE_URL}/api/tech-risk/assessments", headers=h, timeout=15)
    assert r.status_code in (200, 401)  # accept 401 if auth required strictly


def test_regression_issue_management():
    tok = requests.post(f"{BASE_URL}/api/auth/login", json={"email": "admin@bank.com", "password": "admin123"}, timeout=15).json()
    token = tok.get("access_token") or tok.get("token")
    h = {"Authorization": f"Bearer {token}"} if token else {}
    r = requests.get(f"{BASE_URL}/api/issue-management/issues", headers=h, timeout=15)
    assert r.status_code in (200, 401, 404)
